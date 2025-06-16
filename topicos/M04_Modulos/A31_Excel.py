from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

## CArega o arquivo
workbook:Workbook = load_workbook(WORKBOOK_PATH)

sheet_name = 'Planilha1'
# define a abinha ativa
worksheet: Worksheet = workbook[sheet_name]

row = tuple[Cell]

for row in worksheet.iter_rows():
    for cell in row:
        cell.value
        if cell.value == 'pedro':
            worksheet.cell(cell.row, 2, 'nheco')

workbook.save(WORKBOOK_PATH)

