from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()

print(workbook.sheetnames)
# cria abinha
sheet_name = 'Planilha1'
workbook.create_sheet(sheet_name)
print(workbook.sheetnames)

# define a abinha ativa
worksheet: Worksheet = workbook[sheet_name]


worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    # nome     idade  nota
    ['joao',     15,   5.9],
    ['joao',     15,   5.9],
    ['maria',    14,   8.7],
    ['carlos',   16,   4.2],
    ['ana',      15,   9.1],
    ['pedro',    14,   6.5],
    ['lucas',    16,   7.3],
]

# # Totalmente controlavel
# for i, student_row in enumerate(students, start=2):
#     for j , student_column in enumerate(student_row, start=1):
#         # print(i, j, student_column)
#         worksheet.cell(i, j, student_column)

## Mesma coisa porem sem ito controle
for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)

